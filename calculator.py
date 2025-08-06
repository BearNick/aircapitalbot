# Финализированный файл calculator.py со стилями, форматами чисел и улучшенным внешним видом

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime

def generate_financial_model(data: dict, output_dir="output", return_metrics=False):
    # Очистка и парсинг входных данных
    years = int(data["horizon"])
    revenue_year1 = float(str(data["revenue_year1"]).replace(" ", "").replace(",", ""))
    growth_rate = float(str(data["growth"]).replace(",", ".")) / 100
    investment = float(str(data["investment"]).replace(" ", "").replace(",", ""))
    fixed_costs_monthly = float(str(data["fixed_costs"]).replace(" ", "").replace(",", ""))
    variable_costs_pct = float(str(data["variable_costs"]).replace(" ", "").replace(",", "")) / 100
    employees = int(data["employees"])
    avg_salary = float(str(data["avg_salary"]).replace(" ", "").replace(",", ""))

    # Базовые параметры
    tax_rate = 0.20
    discount_rate = 0.12
    salary_costs = avg_salary * employees * 12

    results = []
    cash_flows = []
    cumulative_cash_flow = 0

    for year in range(1, years + 1):
        revenue = revenue_year1 * ((1 + growth_rate) ** (year - 1))
        variable_costs = revenue * variable_costs_pct
        fixed_costs = fixed_costs_monthly * 12
        total_operating_expenses = variable_costs + fixed_costs + salary_costs
        ebitda = revenue - total_operating_expenses
        tax = tax_rate * ebitda if ebitda > 0 else 0
        net_income = ebitda - tax
        discounted_cash_flow = net_income / ((1 + discount_rate) ** year)
        cumulative_cash_flow += net_income

        results.append({
            "Год": year,
            "Выручка (₽)": round(revenue),
            "Переменные расходы (₽)": round(variable_costs),
            "Постоянные расходы (₽)": round(fixed_costs),
            "ФОТ (₽)": round(salary_costs),
            "EBITDA (₽)": round(ebitda),
            "Налог (₽)": round(tax),
            "Чистая прибыль (₽)": round(net_income),
            "DCF (₽)": round(discounted_cash_flow),
            "Кум. прибыль (₽)": round(cumulative_cash_flow)
        })

        cash_flows.append(net_income)

    df_financials = pd.DataFrame(results)

    # Метрики
    npv = sum([cf / ((1 + discount_rate) ** (i + 1)) for i, cf in enumerate(cash_flows)])
    try:
        irr = np.irr([-investment] + cash_flows)
        irr_value = round(irr * 100, 2)
    except:
        irr_value = "N/A"
    payback_year = next((i + 1 for i, s in enumerate(np.cumsum(cash_flows)) if s >= investment), "За пределами горизонта")

    # Excel файл
    os.makedirs(output_dir, exist_ok=True)
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"financial_model_{now}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb = Workbook()

    # Стиль
    bold_center = NamedStyle(name="bold_center")
    bold_center.font = Font(bold=True)
    bold_center.alignment = Alignment(horizontal="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    def apply_styles(ws):
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(name="Calibri", size=11)
                cell.border = border

    # Summary
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Показатель", "Значение"])
    summary_data = [
        ("Общие инвестиции", f"{int(investment):,} ₽"),
        ("NPV", f"{int(npv):,} ₽"),
        ("IRR", f"{irr_value} %"),
        ("Срок окупаемости", f"{payback_year} лет"),
        ("Горизонт планирования", f"{years} лет"),
        ("Ставка дисконтирования", f"{discount_rate * 100:.1f} %"),
        ("Налоговая ставка", f"{tax_rate * 100:.1f} %")
    ]
    for row in summary_data:
        summary.append(row)
    apply_styles(summary)

    # Assumptions
    assumptions = wb.create_sheet("Assumptions")
    assumptions.append(["Параметр", "Значение"])
    assumptions_data = [
        ("Тип проекта", data.get("project_type", "")),
        ("Регион", data.get("region", "")),
        ("Инвестиции", f"{int(investment):,} ₽"),
        ("Выручка в 1-й год", f"{int(revenue_year1):,} ₽"),
        ("Рост выручки", f"{growth_rate * 100:.1f} %"),
        ("Переменные расходы", f"{variable_costs_pct * 100:.1f} % от выручки"),
        ("Постоянные расходы", f"{int(fixed_costs_monthly):,} ₽ в месяц"),
        ("Количество сотрудников", f"{employees} чел."),
        ("Средняя зарплата", f"{int(avg_salary):,} ₽ в месяц")
    ]
    for row in assumptions_data:
        assumptions.append(row)
    apply_styles(assumptions)

    # P&L
    pnl = wb.create_sheet("P&L")
    for r in dataframe_to_rows(df_financials, index=False, header=True):
        pnl.append(r)
    apply_styles(pnl)

    # Multiples
    multiples = wb.create_sheet("Multiples")
    revenue_total = df_financials["Выручка (₽)"].sum()
    ebitda_total = df_financials["EBITDA (₽)"].sum()
    valuation = investment + npv

    multiples.append(["Мультипликатор", "Значение"])
    multiples_data = [
        ("EV/Revenue", round(valuation / revenue_total, 2) if revenue_total else "N/A"),
        ("EV/EBITDA", round(valuation / ebitda_total, 2) if ebitda_total else "N/A")
    ]
    for row in multiples_data:
        multiples.append(row)
    apply_styles(multiples)

    # Справка
    info = wb.create_sheet("Справка")
    help_text = [
        ("EBITDA", "Прибыль до вычета процентов, налогов и амортизации"),
        ("NPV", "Чистая приведённая стоимость — текущая стоимость будущих денежных потоков"),
        ("IRR", "Внутренняя норма доходности — ставка, при которой NPV = 0"),
        ("Payback", "Период времени, за который инвестиции окупаются"),
        ("DCF", "Дисконтированный денежный поток — чистая прибыль с учётом стоимости денег во времени"),
        ("EV/EBITDA", "Оценка компании к прибыли до налогов и амортизации"),
        ("EV/Revenue", "Оценка компании к выручке")
    ]
    info.append(["Показатель", "Описание"])
    for row in help_text:
        info.append(row)
    apply_styles(info)

    wb.save(filepath)

    if return_metrics:
        metrics = {
            "NPV": round(npv, 2),
            "IRR": irr_value,
            "Payback": payback_year
        }
        return filepath, metrics
    else:
        return filepath
def extract_metrics_from_results(metrics: dict) -> str:
    return (
        f"📊 Основные показатели:\n"
        f"• NPV: {metrics.get('NPV', '—')} ₽\n"
        f"• IRR: {metrics.get('IRR', '—')} %\n"
        f"• Срок окупаемости: {metrics.get('Payback', '—')} лет"
    )
